"""
API Endpoints for Legal, IT, Finance, and Sales domain operations.
"""

import io
import json
from pathlib import Path
from typing import Dict, Any, Optional, List
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import PlainTextResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_active_user
from app.models.models import AgentWorkflow, User, WorkflowApproval
from app.services.document_parser import DocumentParseError, extract_file_text
from app.services.legal_service import (
    audit_contract_text,
    check_software_licenses,
    compare_contract_texts,
    detect_sensitive_data,
)
from app.services.contract_review import detect_contract_type, review_contract
from app.services.legal_document_generator import generate_legal_document
from app.services.legal_documents import list_document_schemas, validate_document_fields
from app.services.rag_service import hybrid_search_documents
from app.services.it_service import handle_it_request
from app.services.finance_service import audit_invoice_and_reconcile
from app.services.sales_service import handle_sales_request

router = APIRouter(tags=["Specialized Domain APIs"])
MAX_LEGAL_FILE_BYTES = 10 * 1024 * 1024


class ContractAuditRequest(BaseModel):
    document_name: str = "Contract.pdf"
    contract_text: str
    represented_party: str = "NEUTRAL"


class CreateJiraTicketRequest(BaseModel):
    summary: str
    description: Optional[str] = None
    priority: str = "MEDIUM"


class AuditInvoiceRequest(BaseModel):
    invoice_text: str


class SalesQuotationRequest(BaseModel):
    customer_name: str = "Khách Hàng Doanh Nghiệp"
    item_query: str


class LegalDocumentGenerateRequest(BaseModel):
    document_type: str
    output_format: str = "docx"
    fields: dict[str, Any]


class LegalDocumentValidationRequest(BaseModel):
    document_type: str
    fields: dict[str, Any]


async def _read_legal_file(file: UploadFile) -> tuple[str, str, list[str]]:
    data = await file.read()
    if not data:
        raise HTTPException(status_code=422, detail="The uploaded file is empty")
    if len(data) > MAX_LEGAL_FILE_BYTES:
        raise HTTPException(status_code=413, detail="Legal files are limited to 10 MB")
    filename = Path(file.filename or "document.txt").name
    extension = Path(filename).suffix.lower()
    headers: list[str] = []
    try:
        if extension == ".json":
            payload = json.loads(data.decode("utf-8-sig"))
            text = json.dumps(payload, ensure_ascii=False, indent=2)
            headers = list(payload) if isinstance(payload, dict) else []
        elif extension == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets:
                rows.append(f"# {sheet.title}")
                for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                    values = ["" if value is None else str(value) for value in row]
                    if row_index == 0:
                        headers.extend(value for value in values if value)
                    rows.append(" | ".join(values))
            text = "\n".join(rows)
        else:
            text = extract_file_text(filename, data)
            if extension == ".csv" and text:
                headers = [part.strip() for part in text.splitlines()[0].split("|")]
    except (DocumentParseError, UnicodeDecodeError, ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if not text.strip():
        raise HTTPException(status_code=422, detail="No readable text was found in the file")
    return filename, text, headers


def _create_legal_approval(
    db: Session,
    current_user: User,
    result: dict[str, Any],
    action_type: str = "LEGAL_CONTRACT_APPROVAL",
) -> str | None:
    if not result.get(
        "requires_legal_approval",
        result.get("risk_level") in {"HIGH", "CRITICAL"},
    ):
        return None
    document_name = result.get("document_name") or result.get("manifest") or "Legal review"
    findings = result.get("risks") or result.get("findings") or []
    reason_by_action = {
        "LEGAL_CONTRACT_APPROVAL": "Legal Agent detected high-risk contract terms.",
        "LEGAL_PRIVACY_APPROVAL": "Sensitive or restricted personal data was detected.",
        "LEGAL_LICENSE_APPROVAL": "A reciprocal open-source license requires commercial-use review.",
    }
    workflow = AgentWorkflow(
        tenant_id=current_user.tenant_id,
        initiator_id=current_user.id,
        title=f"Legal review: {document_name}",
        status="AWAITING_APPROVAL",
        current_step=1,
        dag_plan={
            "agent_role": "LEGAL",
            "steps": ["EMPLOYEE_SUBMISSION", "MANAGER_REVIEW", "LEGAL_APPROVAL"],
        },
    )
    db.add(workflow)
    db.flush()
    approval = WorkflowApproval(
        workflow_id=workflow.id,
        action_type=action_type,
        risk_level=result.get("risk_level", "HIGH"),
        payload={
            "document_name": document_name,
            "risk_score": result.get("risk_score"),
            "findings": findings,
            "reason": reason_by_action.get(action_type, "Legal review is required."),
            "requester_name": current_user.full_name,
            "data_sources": [document_name],
        },
        status="WAITING",
    )
    db.add(approval)
    db.commit()
    return str(workflow.id)


# --- LEGAL ---
def _retrieve_contract_review_references(
    db: Session,
    current_user: User,
    contract_type_label: str,
) -> list[dict[str, Any]]:
    """Retrieve ACL-filtered internal policy/template context for Legal review."""
    try:
        chunks = hybrid_search_documents(
            db=db,
            tenant_id=current_user.tenant_id,
            query_text=(
                f"{contract_type_label} mẫu hợp đồng chuẩn policy pháp lý "
                "thanh toán trách nhiệm sở hữu trí tuệ chấm dứt bảo mật"
            ),
            department=(
                "*"
                if current_user.role in {"Owner", "Admin", "CEO"}
                else current_user.department
            ),
            top_k=6,
            user_role=current_user.role,
            user_department=current_user.department,
        )
    except Exception:
        # Contract analysis still works deterministically if the tenant has no
        # indexed legal knowledge or its retrieval service is temporarily down.
        db.rollback()
        return []

    references: list[dict[str, Any]] = []
    seen_documents: set[tuple[str, str]] = set()
    for chunk in chunks:
        name = str(chunk.get("document_title") or chunk.get("document_name") or "Tài liệu nội bộ")
        document_id = str(chunk.get("document_id") or chunk.get("document_name") or "")
        version = str(chunk.get("version") or "1.0")
        document_key = (document_id, version)
        if not document_id or document_key in seen_documents:
            continue
        seen_documents.add(document_key)
        normalized_name = name.lower()
        source_type = (
            "APPROVED_TEMPLATE"
            if any(token in normalized_name for token in ("template", "mẫu", "mau", "approved", "chuẩn"))
            else "COMPANY_POLICY"
        )
        references.append({
            "id": document_id,
            "document_id": document_id,
            "version": version,
            "type": source_type,
            "title": name,
            "section_title": chunk.get("section_title"),
            "citation_tag": chunk.get("citation_tag"),
            "score": chunk.get("score"),
            "url": "",
            "reader_url": (
                f"/api/v1/documents/{quote(document_id, safe='')}/reader"
                f"?version={quote(version, safe='')}"
            ),
            "note": "Ngữ cảnh nội bộ đã truy xuất theo ACL; Legal cần xác nhận tính áp dụng.",
        })
    return references


@router.get("/legal/document-templates", summary="List schema-driven legal document templates")
def list_legal_document_templates(
    current_user: User = Depends(get_current_active_user),
) -> List[Dict[str, Any]]:
    return list_document_schemas()


@router.post("/legal/validate-document", summary="Validate a legal document draft before generation")
def validate_legal_document_endpoint(
    req: LegalDocumentValidationRequest,
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    try:
        return validate_document_fields(req.document_type, req.fields)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@router.post("/legal/audit-contract", summary="Audit contract text for high-risk clauses")
def audit_contract_endpoint(
    req: ContractAuditRequest,
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    try:
        return review_contract(
            req.contract_text,
            req.document_name,
            req.represented_party,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@router.post("/legal/review-document", summary="Extract and review a legal document")
async def review_legal_document(
    file: UploadFile = File(...),
    represented_party: str = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    filename, text, _ = await _read_legal_file(file)
    detection = detect_contract_type(text)
    references = _retrieve_contract_review_references(
        db, current_user, detection["contract_type_label"]
    )
    try:
        result = review_contract(text, filename, represented_party, references)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    result["workflow_id"] = _create_legal_approval(db, current_user, result)
    result["approval_created"] = result["workflow_id"] is not None
    return result


@router.post("/legal/compare-documents", summary="Compare two contract versions")
async def compare_legal_documents(
    old_file: UploadFile = File(...),
    new_file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    old_name, old_text, _ = await _read_legal_file(old_file)
    new_name, new_text, _ = await _read_legal_file(new_file)
    result = compare_contract_texts(old_text, new_text)
    return {"old_document": old_name, "new_document": new_name, **result}


@router.post("/legal/privacy-check", summary="Detect personal and restricted data")
async def privacy_check_document(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    filename, text, headers = await _read_legal_file(file)
    result = {"document_name": filename, **detect_sensitive_data(text, headers)}
    result["workflow_id"] = _create_legal_approval(
        db, current_user, result, "LEGAL_PRIVACY_APPROVAL"
    )
    result["approval_created"] = result["workflow_id"] is not None
    return result


@router.post("/legal/license-check", summary="Inspect a software dependency manifest")
async def license_check_manifest(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    filename, text, _ = await _read_legal_file(file)
    result = check_software_licenses(filename, text)
    result["workflow_id"] = _create_legal_approval(
        db, current_user, result, "LEGAL_LICENSE_APPROVAL"
    )
    result["approval_created"] = result["workflow_id"] is not None
    return result


@router.post("/legal/generate-document", summary="Generate an editable legal document draft")
def generate_legal_document_endpoint(
    req: LegalDocumentGenerateRequest,
    current_user: User = Depends(get_current_active_user),
):
    try:
        content, filename, media_type = generate_legal_document(
            req.document_type, req.output_format, req.fields
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/legal/download-redline/{file_id}", response_class=PlainTextResponse, summary="Download contract redline docx file")
def download_redline(file_id: str):
    return f"SIMULATED REDLINE DOCX FILE FOR {file_id}\nAll penalty clauses adjusted to 8% max limit per Law."


# --- IT ---
@router.post("/it/tickets", summary="Create Jira ticket for technical issue")
def create_jira_ticket_endpoint(
    req: CreateJiraTicketRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    return handle_it_request(db, current_user, f"{req.summary} - {req.description or ''}")


# --- FINANCE ---
@router.post("/finance/audit-invoice", summary="Audit invoice text and reconcile PO")
def audit_invoice_endpoint(
    req: AuditInvoiceRequest,
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    return audit_invoice_and_reconcile(req.invoice_text)


# --- SALES ---
@router.post("/sales/quotation", summary="Generate sales quotation PDF payload")
def generate_quotation_endpoint(
    req: SalesQuotationRequest,
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    return handle_sales_request(req.item_query, customer_name=req.customer_name)


@router.get("/sales/download-quote/{file_id}", response_class=PlainTextResponse, summary="Download sales PDF quotation file")
def download_quote(file_id: str):
    return f"SIMULATED PDF QUOTATION FILE FOR {file_id}\nOfficial AI Workforce Quotation Document."
