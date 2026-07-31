"""
Dashboard & Analytics API router — Returns real statistics from Database.
"""

from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from typing import Dict, Any, List
from io import BytesIO

from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.database import get_db
from app.core.security import get_current_active_user, get_current_user
from app.models.models import (
    User,
    AIAgent,
    AgentWorkflow,
    WorkflowApproval,
    AuditLog,
    ChatConversation,
    ChatMessage,
)

from app.services.auth_service import ensure_tenant_default_agents

router = APIRouter(prefix="/dashboard", tags=["Dashboard"])

ROLE_DEPT_MAP = {
    "CEO": "Board of Directors",
    "HR": "Human Resources",
    "KNOWLEDGE": "Knowledge Base",
    "LEGAL": "Legal Department",
    "IT": "IT Department",
    "FINANCE": "Finance & Accounting",
    "SALES": "Sales & Marketing",
}

DASHBOARD_TIMEZONE = ZoneInfo("Asia/Ho_Chi_Minh")


def _period_bounds(period: str, now: datetime | None = None) -> tuple[datetime, datetime, datetime]:
    """Return selected-period UTC bounds and the current local dashboard time."""
    now_utc = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
    now_local = now_utc.astimezone(DASHBOARD_TIMEZONE)
    today_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0)

    if period == "day":
        start_local = today_local
    elif period == "week":
        start_local = today_local - timedelta(days=6)
    else:
        start_local = today_local.replace(day=1)

    return start_local.astimezone(timezone.utc), now_utc, now_local


def _month_start(value: datetime, offset: int = 0) -> datetime:
    """Shift an aware datetime to the first day of a calendar month."""
    month_index = value.year * 12 + value.month - 1 + offset
    year, zero_based_month = divmod(month_index, 12)
    return value.replace(
        year=year,
        month=zero_based_month + 1,
        day=1,
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
    )


def _build_usage_trend(
    db: Session,
    tenant_id,
    period: str,
    period_start: datetime,
    period_end: datetime,
    now_local: datetime,
) -> List[Dict[str, Any]]:
    """Build period-aware usage buckets in Vietnam local time."""
    assistant_times = [
        row[0]
        for row in db.query(ChatMessage.created_at)
        .join(ChatConversation, ChatMessage.conversation_id == ChatConversation.id)
        .filter(
            ChatConversation.tenant_id == tenant_id,
            ChatMessage.sender == "ASSISTANT",
            ChatMessage.created_at >= period_start,
            ChatMessage.created_at <= period_end,
        )
        .all()
    ]
    handoff_times = [
        row[0]
        for row in db.query(WorkflowApproval.updated_at)
        .join(AgentWorkflow)
        .filter(
            AgentWorkflow.tenant_id == tenant_id,
            WorkflowApproval.updated_at >= period_start,
            WorkflowApproval.updated_at <= period_end,
        )
        .all()
    ]

    if period == "day":
        keys = list(range(24))
        labels = {hour: f"{hour:02d}h" for hour in keys}
        key_for = lambda value: value.astimezone(DASHBOARD_TIMEZONE).hour
    else:
        first_local = period_start.astimezone(DASHBOARD_TIMEZONE)
        day_count = (now_local.date() - first_local.date()).days + 1
        local_days = [first_local.date() + timedelta(days=i) for i in range(day_count)]
        keys = local_days
        if period == "week":
            day_labels = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]
            labels = {day: day_labels[day.weekday()] for day in keys}
        else:
            labels = {day: day.strftime("%d/%m") for day in keys}
        key_for = lambda value: value.astimezone(DASHBOARD_TIMEZONE).date()

    ai_counts = {key: 0 for key in keys}
    handoff_counts = {key: 0 for key in keys}
    for created_at in assistant_times:
        key = key_for(created_at)
        if key in ai_counts:
            ai_counts[key] += 1
    for updated_at in handoff_times:
        key = key_for(updated_at)
        if key in handoff_counts:
            handoff_counts[key] += 1

    return [
        {
            "day": labels[key],
            "ai": ai_counts[key],
            "handoff": handoff_counts[key],
        }
        for key in keys
    ]


@router.get("/stats", summary="Get real dashboard statistics from Database")
def get_dashboard_stats(
    period: str = Query("week", enum=["day", "week", "month"]),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Dict[str, Any]:
    tenant_id = current_user.tenant_id
    period_start, period_end, now_local = _period_bounds(period)

    # 1. AI Agents Stats (Auto-seed if missing)
    agents = ensure_tenant_default_agents(db, tenant_id)
    total_agents = len(agents)
    active_agents = sum(1 for a in agents if a.is_active)

    # 2. Employees Stats
    total_employees = db.query(User).filter(User.tenant_id == tenant_id).count()
    active_employees = db.query(User).filter(User.tenant_id == tenant_id, User.is_active == True).count()

    # 3. Total chat messages. Audit logs and workflows are operational events,
    # not user/assistant messages, so they must not be used for this KPI.
    message_query = db.query(ChatMessage).join(
        ChatConversation,
        ChatMessage.conversation_id == ChatConversation.id,
    ).filter(
        ChatConversation.tenant_id == tenant_id,
        ChatMessage.created_at >= period_start,
        ChatMessage.created_at <= period_end,
    )
    total_messages = message_query.count()
    total_workflows = db.query(AgentWorkflow).filter(
        AgentWorkflow.tenant_id == tenant_id,
        AgentWorkflow.created_at >= period_start,
        AgentWorkflow.created_at <= period_end,
    ).count()

    # 4. Handoffs (Approvals)
    total_approvals = db.query(WorkflowApproval).join(AgentWorkflow).filter(
        AgentWorkflow.tenant_id == tenant_id,
        WorkflowApproval.updated_at >= period_start,
        WorkflowApproval.updated_at <= period_end,
    ).count()
    handoff_rate = round((total_approvals / (total_workflows or 1)) * 100, 1)

    # 5. Chatbots Table Data with Dynamic Accuracy from Audit Logs
    chatbots_data = []
    for agent in agents:
        agent_logs = db.query(AuditLog).filter(
            AuditLog.tenant_id == tenant_id,
            AuditLog.agent_role == agent.role_code,
            AuditLog.created_at >= period_start,
            AuditLog.created_at <= period_end,
        ).all()
        
        agent_logs_count = len(agent_logs)
        conversation_count = db.query(func.count(func.distinct(ChatMessage.conversation_id))).join(
            ChatConversation,
            ChatMessage.conversation_id == ChatConversation.id,
        ).filter(
            ChatConversation.tenant_id == tenant_id,
            ChatConversation.ai_agent_id == agent.id,
            ChatMessage.created_at >= period_start,
            ChatMessage.created_at <= period_end,
        ).scalar() or 0
        
        # Calculate real Accuracy: ratio of non-error execution outputs or default 98.0% if active
        if not agent.is_active:
            calc_accuracy = 0.0
        elif agent_logs_count > 0:
            error_count = sum(1 for log in agent_logs if log.output_result and "error" in log.output_result)
            calc_accuracy = round(((agent_logs_count - error_count) / agent_logs_count) * 100, 1)
        else:
            calc_accuracy = 98.0  # Default baseline for active agents
        
        chatbots_data.append({
            "id": str(agent.id),
            "name": agent.name,
            "emoji": agent.avatar_emoji or "🤖",
            "dept": ROLE_DEPT_MAP.get(agent.role_code, agent.role_code),
            "role_code": agent.role_code,
            "conversations": conversation_count,
            "accuracy": calc_accuracy,
            "status": "active" if agent.is_active else "inactive",
        })

    # 6. Top Employee Users ordered by real USER messages in the selected period.
    message_counts = db.query(
        ChatConversation.user_id.label("user_id"),
        func.count(ChatMessage.id).label("message_count"),
    ).join(
        ChatMessage,
        ChatMessage.conversation_id == ChatConversation.id,
    ).filter(
        ChatConversation.tenant_id == tenant_id,
        ChatMessage.sender == "USER",
        ChatMessage.created_at >= period_start,
        ChatMessage.created_at <= period_end,
    ).group_by(ChatConversation.user_id).subquery()

    count_value = func.coalesce(message_counts.c.message_count, 0)
    top_users = db.query(User, count_value.label("message_count")).outerjoin(
        message_counts,
        message_counts.c.user_id == User.id,
    ).filter(
        User.tenant_id == tenant_id,
    ).order_by(
        count_value.desc(),
        User.full_name.asc(),
    ).limit(5).all()

    max_user_messages = max((int(count) for _, count in top_users), default=0)
    top_employees_data = []
    for u, user_messages_count in top_users:
        user_messages_count = int(user_messages_count)
        # Initials for avatar
        name_parts = u.full_name.split()
        initials = "".join([p[0] for p in name_parts[:2]]).upper() if name_parts else "U"
        
        top_employees_data.append({
            "id": str(u.id),
            "name": u.full_name,
            "dept": u.department,
            "avatar": initials,
            "msgs": user_messages_count,
            "pct": round((user_messages_count / max_user_messages) * 100) if max_user_messages else 0,
        })

    # 7. Usage Trend Data follows the selected day/week/month period.
    usage_trend = _build_usage_trend(
        db, tenant_id, period, period_start, period_end, now_local
    )

    # 8. Monthly Data: all 12 months of the current local calendar year.
    monthly_data = []
    current_month_local = _month_start(now_local)
    year_start_local = current_month_local.replace(month=1)
    for offset in range(12):
        month_start_local = _month_start(year_start_local, offset)
        next_month_local = _month_start(year_start_local, offset + 1)
        month_messages = db.query(ChatMessage).join(
            ChatConversation,
            ChatMessage.conversation_id == ChatConversation.id,
        ).filter(
            ChatConversation.tenant_id == tenant_id,
            ChatMessage.created_at >= month_start_local.astimezone(timezone.utc),
            ChatMessage.created_at < next_month_local.astimezone(timezone.utc),
        ).count()
        monthly_data.append({
            "month": f"T{month_start_local.month}",
            "value": month_messages,
        })

    # Calculated Monthly Mini Stats
    current_month_messages = db.query(ChatMessage).join(
        ChatConversation,
        ChatMessage.conversation_id == ChatConversation.id,
    ).filter(
        ChatConversation.tenant_id == tenant_id,
        ChatMessage.created_at >= current_month_local.astimezone(timezone.utc),
        ChatMessage.created_at <= period_end,
    ).count()
    
    avg_per_day = round(current_month_messages / max(now_local.day, 1), 1)

    return {
        "kpi": {
            "active_bots": f"{active_agents} / {total_agents}",
            "total_agents": total_agents,
            "active_agents": active_agents,
            "total_messages": total_messages,
            "active_employees": active_employees,
            "total_employees": total_employees,
            "handoff_rate": f"{handoff_rate}%",
        },
        "monthly_summary": {
            "target": 500,
            "completed": total_messages,
            "avg_per_day": avg_per_day,
        },
        "chatbots": chatbots_data,
        "top_employees": top_employees_data,
        "usage_trend": usage_trend,
        "monthly_data": monthly_data,
        "period": {
            "key": period,
            "start": period_start.isoformat(),
            "end": period_end.isoformat(),
            "timezone": str(DASHBOARD_TIMEZONE),
        },
    }


# ── Report Export Endpoints ──────────────────────────────────────────────────
@router.get("/reports/export/excel", summary="Export Dashboard Report to Excel")
def export_report_excel(
    period: str = Query("week"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    stats = get_dashboard_stats(period=period, db=db, current_user=current_user)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Report"

        # Title Banner
        ws.merge_cells("A1:F1")
        ws["A1"] = "AI WORKFORCE — EXECUTIVE DASHBOARD REPORT"
        ws["A1"].font = Font(name="Calibri", size=15, bold=True, color="3C50E0")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        ws["A2"] = f"Exported At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Period: {period.upper()}"
        ws["A2"].font = Font(italic=True, color="64748B")

        header_fill = PatternFill(start_color="3C50E0", end_color="3C50E0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # 1. KPI Summary
        ws["A4"] = "1. KEY PERFORMANCE INDICATORS"
        ws["A4"].font = Font(bold=True, size=11, color="1E293B")

        headers_kpi = ["Active Chatbots", "Total Messages", "Active Employees", "Handoff Rate"]
        values_kpi = [
            stats["kpi"]["active_bots"],
            stats["kpi"]["total_messages"],
            f"{stats['kpi']['active_employees']} / {stats['kpi']['total_employees']}",
            stats["kpi"]["handoff_rate"]
        ]

        for col_num, h_text in enumerate(headers_kpi, 1):
            cell = ws.cell(row=5, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_num, val in enumerate(values_kpi, 1):
            cell = ws.cell(row=6, column=col_num, value=val)
            cell.alignment = Alignment(horizontal="center")

        # 2. Active Chatbots Overview
        ws["A9"] = "2. ACTIVE CHATBOTS OVERVIEW"
        ws["A9"].font = Font(bold=True, size=11, color="1E293B")

        headers_bots = ["Name", "Department", "Role Code", "Conversations", "Accuracy", "Status"]
        for col_num, h_text in enumerate(headers_bots, 1):
            cell = ws.cell(row=10, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        row_idx = 11
        for bot in stats["chatbots"]:
            ws.cell(row=row_idx, column=1, value=bot["name"])
            ws.cell(row=row_idx, column=2, value=bot["dept"])
            ws.cell(row=row_idx, column=3, value=bot["role_code"])
            ws.cell(row=row_idx, column=4, value=bot["conversations"]).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=5, value=f"{bot['accuracy']}%").alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=6, value=bot["status"].upper()).alignment = Alignment(horizontal="center")
            row_idx += 1

        # 3. Top Employees Overview
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="3. TOP EMPLOYEES OVERVIEW").font = Font(bold=True, size=11, color="1E293B")
        row_idx += 1
        headers_emp = ["Name", "Department", "Messages Executed"]
        for col_num, h_text in enumerate(headers_emp, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1
        for emp in stats["top_employees"]:
            ws.cell(row=row_idx, column=1, value=emp["name"])
            ws.cell(row=row_idx, column=2, value=emp["dept"])
            ws.cell(row=row_idx, column=3, value=emp["msgs"]).alignment = Alignment(horizontal="right")
            row_idx += 1

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=chatbot-report-{period}.xlsx"}
        )
    except Exception as e:
        csv_lines = [
            "AI WORKFORCE - DASHBOARD REPORT",
            f"Exported At,{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Tenant ID,{current_user.tenant_id}",
            "",
            "--- KPI SUMMARY ---",
            f"Active Chatbots,{stats['kpi']['active_bots']}",
            f"Total Messages,{stats['kpi']['total_messages']}",
            f"Active Employees,{stats['kpi']['active_employees']}",
            f"Handoff Rate,{stats['kpi']['handoff_rate']}",
            "",
            "--- CHATBOTS OVERVIEW ---",
            "Name,Department,Role,Conversations,Accuracy,Status",
        ]
        for bot in stats["chatbots"]:
            csv_lines.append(f"\"{bot['name']}\",\"{bot['dept']}\",{bot['role_code']},{bot['conversations']},{bot['accuracy']}%,{bot['status']}")
        csv_lines.append("")
        csv_lines.append("--- EMPLOYEES OVERVIEW ---")
        csv_lines.append("Name,Department,Messages Executed")
        for emp in stats["top_employees"]:
            csv_lines.append(f"\"{emp['name']}\",\"{emp['dept']}\",{emp['msgs']}")

        csv_content = "\n".join(csv_lines)
        return Response(
            content=csv_content.encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=chatbot-report-{period}.csv"}
        )


@router.get("/reports/export/pdf", summary="Export Dashboard Report to PDF")
def export_report_pdf(
    period: str = Query("week"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    stats = get_dashboard_stats(period=period, db=db, current_user=current_user)
    
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=16,
            leading=20,
            textColor=colors.HexColor('#3C50E0'),
            spaceAfter=4
        )
        subtitle_style = ParagraphStyle(
            'ReportSubTitle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=12
        )
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=11,
            leading=14,
            textColor=colors.HexColor('#1E293B'),
            spaceBefore=10,
            spaceAfter=6
        )

        elements = []
        
        # Header
        elements.append(Paragraph(f"AI WORKFORCE — EXECUTIVE REPORT ({period.upper()})", title_style))
        elements.append(Paragraph(f"Generated for: {current_user.full_name} ({current_user.email}) | Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
        elements.append(Spacer(1, 8))

        # KPI Table
        elements.append(Paragraph("1. KEY PERFORMANCE INDICATORS", section_style))
        kpi_data = [
            ["Active Chatbots", "Total Messages", "Active Employees", "Handoff Rate"],
            [
                stats['kpi']['active_bots'],
                str(stats['kpi']['total_messages']),
                f"{stats['kpi']['active_employees']} / {stats['kpi']['total_employees']}",
                stats['kpi']['handoff_rate']
            ]
        ]
        t_kpi = Table(kpi_data, colWidths=[130, 130, 140, 140])
        t_kpi.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3C50E0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(t_kpi)
        elements.append(Spacer(1, 12))

        # Chatbots Table
        elements.append(Paragraph("2. ACTIVE CHATBOTS STATUS", section_style))
        bot_table_data = [["Name", "Department", "Role", "Msgs", "Accuracy", "Status"]]
        for bot in stats["chatbots"]:
            bot_table_data.append([
                bot['name'],
                bot['dept'],
                bot['role_code'],
                str(bot['conversations']),
                f"{bot['accuracy']}%",
                bot['status'].upper()
            ])
        t_bots = Table(bot_table_data, colWidths=[130, 120, 80, 60, 80, 70])
        t_bots.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (3, 1), (4, -1), 'RIGHT'),
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(t_bots)
        elements.append(Spacer(1, 12))

        # Top Employees Table
        elements.append(Paragraph("3. TOP EMPLOYEE USERS", section_style))
        emp_table_data = [["Name", "Department", "Workflows Executed"]]
        for emp in stats["top_employees"]:
            emp_table_data.append([
                emp['name'],
                emp['dept'],
                str(emp['msgs'])
            ])
        t_emp = Table(emp_table_data, colWidths=[220, 180, 140])
        t_emp.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#475569')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(t_emp)

        doc.build(elements)
        buffer.seek(0)
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=chatbot-report-{period}.pdf"}
        )
    except Exception as e:
        report_text = f"AI WORKFORCE PLATFORM REPORT ({period.upper()})\n\nKPIs:\nActive Bots: {stats['kpi']['active_bots']}\nTotal Messages: {stats['kpi']['total_messages']}\n"
        return Response(
            content=report_text.encode("utf-8"),
            media_type="text/plain",
            headers={"Content-Disposition": f"attachment; filename=chatbot-report-{period}.txt"}
        )
