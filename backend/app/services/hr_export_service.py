"""Generate tenant-safe HR directory exports in supported file formats."""

from __future__ import annotations

import html
import json
import os
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Literal

from sqlalchemy.orm import Session

from app.models.models import User
from app.services.hr_employee_tools import export_company_users_dataset

HRExportFormat = Literal["xlsx", "pdf", "json"]
HRDirectoryType = Literal["employees", "managers"]

EXPORT_COLUMNS = (
    ("employee_code", "Mã nhân viên"),
    ("name", "Họ và tên"),
    ("email", "Email"),
    ("role", "Vai trò"),
    ("department", "Phòng ban"),
    ("job_title", "Chức danh"),
    ("employment_status", "Trạng thái"),
    ("manager_name", "Quản lý trực tiếp"),
)


def _safe_text(value: Any) -> str:
    return "" if value is None else str(value)


def _spreadsheet_safe(value: Any) -> Any:
    """Prevent exported user-controlled text from becoming an Excel formula."""
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _build_xlsx(items: list[dict[str, Any]], metadata: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Danh sach nhan su"
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False

    last_column = chr(ord("A") + len(EXPORT_COLUMNS) - 1)
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = "AI WORKFORCE - DANH SÁCH NHÂN SỰ"
    sheet["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="3C50E0")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = (
        f"Phạm vi: {metadata['scope']} | Loại: {metadata['directory_type']} | "
        f"Tổng số: {metadata['total_count']} | Xuất lúc: {metadata['generated_at']}"
    )
    sheet["A2"].font = Font(name="Calibri", size=10, italic=True, color="64748B")

    headers = [label for _, label in EXPORT_COLUMNS]
    sheet.append([])
    sheet.append(headers)
    for item in items:
        sheet.append([
            _spreadsheet_safe(_safe_text(item.get(key)))
            for key, _ in EXPORT_COLUMNS
        ])

    header_fill = PatternFill("solid", fgColor="1E293B")
    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = Font(name="Calibri", color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if items:
        table = Table(displayName="HRDirectoryExport", ref=f"A4:{last_column}{4 + len(items)}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    widths = [18, 28, 34, 16, 20, 24, 18, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(ord("A") + index - 1)].width = width
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_font_name() -> str:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = (
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts" / "arial.ttf",
    )
    for candidate in candidates:
        if candidate.is_file():
            pdfmetrics.registerFont(TTFont("HRExportUnicode", str(candidate)))
            return "HRExportUnicode"
    return "Helvetica"


def _build_pdf(items: list[dict[str, Any]], metadata: dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font_name = _pdf_font_name()
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Danh sách nhân sự",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "HRExportTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#3C50E0"),
        alignment=0,
        spaceAfter=5,
    )
    meta_style = ParagraphStyle(
        "HRExportMeta",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#64748B"),
    )
    cell_style = ParagraphStyle(
        "HRExportCell",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=7,
        leading=9,
    )
    header_style = ParagraphStyle(
        "HRExportHeader",
        parent=cell_style,
        textColor=colors.white,
    )

    elements = [
        Paragraph("AI WORKFORCE - DANH SÁCH NHÂN SỰ", title_style),
        Paragraph(
            html.escape(
                f"Phạm vi: {metadata['scope']} | Loại: {metadata['directory_type']} | "
                f"Tổng số: {metadata['total_count']} | Xuất lúc: {metadata['generated_at']}"
            ),
            meta_style,
        ),
        Spacer(1, 8),
    ]
    table_data = [[Paragraph(html.escape(label), header_style) for _, label in EXPORT_COLUMNS]]
    table_data.extend([
        [Paragraph(html.escape(_safe_text(item.get(key))), cell_style) for key, _ in EXPORT_COLUMNS]
        for item in items
    ])
    column_widths = [24, 39, 49, 25, 31, 36, 28, 39]
    table = Table(
        table_data,
        colWidths=[width * mm for width in column_widths],
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
    ]))
    elements.append(table)
    document.build(elements)
    return output.getvalue()


def create_hr_directory_export(
    db: Session,
    *,
    actor: User,
    export_format: HRExportFormat,
    directory_type: HRDirectoryType = "employees",
) -> dict[str, Any]:
    roles = ["Admin", "Manager"] if directory_type == "managers" else None
    dataset = export_company_users_dataset(
        db,
        actor=actor,
        roles=roles,
        active_only=True,
    )
    generated_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    metadata = {
        "request_id": dataset["request_id"],
        "scope": dataset["scope"],
        "directory_type": directory_type,
        "total_count": dataset["total_count"],
        "generated_at": generated_at,
    }
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    base_name = f"hr-{directory_type}-{timestamp}"

    if export_format == "xlsx":
        content = _build_xlsx(dataset["items"], metadata)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif export_format == "pdf":
        content = _build_pdf(dataset["items"], metadata)
        media_type = "application/pdf"
    else:
        payload = {"metadata": metadata, "items": dataset["items"]}
        content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        media_type = "application/json"

    return {
        "content": content,
        "media_type": media_type,
        "filename": f"{base_name}.{export_format}",
        "metadata": metadata,
    }
