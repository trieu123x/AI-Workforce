"""
Tests for HR Agent intent handling, tool execution, leave balances, and Approval Cards.
"""

import json
from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.database import SyncSessionLocal
from app.models.models import AIAgent, AuditLog, User, UserMemory
from app.services.agents.agent_executor import (
    _classify_hr_intent,
    _repair_hr_agent_capabilities,
)


@pytest.mark.parametrize(
    ("message", "expected"),
    [
        ("tìm các nhân viên quản lý", "MANAGER_DIRECTORY"),
        ("tim cac nhan vien quan ly", "MANAGER_DIRECTORY"),
        ("có bao nhiêu quản lý", "MANAGER_DIRECTORY"),
        ("công ty có mấy nhân viên", "EMPLOYEE_DIRECTORY"),
        ("số lượng nhân sự", "EMPLOYEE_DIRECTORY"),
        ("danh sach nhan vien", "EMPLOYEE_DIRECTORY"),
        ("luong cua toi", "SELF_COMPENSATION"),
        ("hop dong sap het han", "CONTRACT_EXPIRY"),
        ("quy định nghỉ phép là gì", "POLICY_QUERY"),
        ("có bao nhiêu nhân viên đang nghỉ phép", "EMPLOYEE_LEAVE_STATUS_COUNT"),
        ("xin chào", "UNKNOWN"),
    ],
)
def test_hr_intent_classifier_normalizes_actions_and_entities(message, expected):
    assert _classify_hr_intent(message) == expected


def test_stale_hr_agent_capabilities_are_split_into_narrow_profile_tools(
    client,
    employee_token_headers,
    transactional_db_session,
):
    agent = transactional_db_session.query(AIAgent).filter(
        AIAgent.role_code == "HR"
    ).first()
    agent.tools_access = [
        "query_leave_balance",
        "request_leave",
        "hybrid_rag_search",
        "get_employee_profile",
    ]
    agent.allowed_actions = list(agent.tools_access)
    agent.disallowed_actions = []
    agent.configuration_version = 1
    _repair_hr_agent_capabilities(agent)
    transactional_db_session.commit()

    assert "get_employee_profile" not in agent.tools_access
    assert "get_employee_basic_profile" in agent.tools_access
    assert "get_employee_full_profile" in agent.tools_access
    assert "query_company_users_sql" in agent.tools_access
    assert "export_hr_directory" in agent.tools_access
    assert agent.configuration_version == 5

    response = client.post(
        "/api/v1/agent/chat",
        json={"agent_role": "HR", "message": "Hồ sơ của tôi"},
        headers=employee_token_headers,
    )
    assert response.status_code == 200, response.text
    assert response.json()["hr_card"]["type"] == "EMPLOYEE_PROFILE"


def test_hr_company_user_sql_tool_respects_chat_actor_scope(
    client,
    employee_token_headers,
    ceo_token_headers,
):
    employee_response = client.post(
        "/api/v1/agent/chat",
        json={"agent_role": "HR", "message": "Danh sách nhân viên"},
        headers=employee_token_headers,
    )
    assert employee_response.status_code == 200, employee_response.text
    employee_data = employee_response.json()
    assert employee_data["tools_executed"][0]["tool_name"] == "query_company_users_sql"
    assert employee_data["hr_card"]["scope"] == "SELF"
    assert len(employee_data["hr_card"]["items"]) == 1
    assert employee_data["hr_card"]["items"][0]["employee"]["email"] == "employee@company.com"

    ceo_response = client.post(
        "/api/v1/agent/chat",
        json={"agent_role": "HR", "message": "Danh sách nhân viên"},
        headers=ceo_token_headers,
    )
    assert ceo_response.status_code == 200, ceo_response.text
    ceo_data = ceo_response.json()
    assert ceo_data["tools_executed"][0]["tool_name"] == "query_company_users_sql"
    assert ceo_data["hr_card"]["scope"] == "COMPANY"
    assert len(ceo_data["hr_card"]["items"]) > 1


def test_hr_manager_directory_phrase_from_chat_routes_to_sql(
    client,
    ceo_token_headers,
):
    response = client.post(
        "/api/v1/agent/chat",
        json={"agent_role": "HR", "message": "tìm các nhân viên quản lý"},
        headers=ceo_token_headers,
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["tools_executed"][0]["tool_name"] == "query_company_users_sql"
    assert data["hr_card"]["directory_type"] == "MANAGERS"
    assert data["hr_card"]["total_count"] == len(data["hr_card"]["items"])
    assert data["hr_card"]["items"]
    assert {
        item["employee"]["role"] for item in data["hr_card"]["items"]
    } <= {"Admin", "Manager"}


def test_hr_unknown_and_unsupported_operational_queries_do_not_fall_back_to_policy(
    client,
    ceo_token_headers,
):
    unknown = client.post(
        "/api/v1/agent/chat",
        json={"agent_role": "HR", "message": "xin chào"},
        headers=ceo_token_headers,
    )
    assert unknown.status_code == 200, unknown.text
    assert unknown.json()["tools_executed"] == []
    assert "chưa xác định rõ" in unknown.json()["reply"].lower()

    unsupported = client.post(
        "/api/v1/agent/chat",
        json={"agent_role": "HR", "message": "có bao nhiêu nhân viên đang nghỉ phép"},
        headers=ceo_token_headers,
    )
    assert unsupported.status_code == 200, unsupported.text
    assert unsupported.json()["tools_executed"] == []
    assert "chưa có tool" in unsupported.json()["reply"].lower()


def test_hr_export_intent_requires_scope_and_format(
    client,
    ceo_token_headers,
):
    for message in ("Xuất file", "Trích xuất file"):
        incomplete = client.post(
            "/api/v1/agent/chat",
            json={"agent_role": "HR", "message": message},
            headers=ceo_token_headers,
        )
        assert incomplete.status_code == 200, incomplete.text
        incomplete_data = incomplete.json()
        assert incomplete_data["tools_executed"] == []
        assert "loại dữ liệu" in incomplete_data["reply"]
        assert "định dạng" in incomplete_data["reply"]

    ready = client.post(
        "/api/v1/agent/chat",
        json={"agent_role": "HR", "message": "Xuất danh sách nhân viên Excel"},
        headers=ceo_token_headers,
    )
    assert ready.status_code == 200, ready.text
    ready_data = ready.json()
    assert ready_data["tools_executed"][0]["tool_name"] == "export_hr_directory"
    assert ready_data["hr_card"]["type"] == "FILE_EXPORT"
    assert ready_data["hr_card"]["format"] == "xlsx"
    assert ready_data["hr_card"]["directory_type"] == "employees"
    assert ready_data["hr_card"]["download_url"].startswith(
        "/api/v1/hr/employees/export?"
    )


def test_hr_directory_export_formats_and_scope(
    client,
    employee_token_headers,
    ceo_token_headers,
    transactional_db_session,
):
    excel = client.get(
        "/api/v1/hr/employees/export?format=xlsx&directory=employees",
        headers=ceo_token_headers,
    )
    assert excel.status_code == 200, excel.text
    assert excel.content.startswith(b"PK")
    assert excel.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(excel.content), read_only=True)
    sheet = workbook["Danh sach nhan su"]
    assert [cell.value for cell in sheet[4]] == [
        "Mã nhân viên",
        "Họ và tên",
        "Email",
        "Vai trò",
        "Phòng ban",
        "Chức danh",
        "Trạng thái",
        "Quản lý trực tiếp",
    ]
    assert sheet.max_row > 4

    pdf = client.get(
        "/api/v1/hr/employees/export?format=pdf&directory=managers",
        headers=ceo_token_headers,
    )
    assert pdf.status_code == 200, pdf.text
    assert pdf.content.startswith(b"%PDF-")
    assert pdf.headers["content-type"].startswith("application/pdf")
    pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf.content)).pages)
    assert "DANH SÁCH NHÂN SỰ" in pdf_text

    own_json = client.get(
        "/api/v1/hr/employees/export?format=json&directory=employees",
        headers=employee_token_headers,
    )
    assert own_json.status_code == 200, own_json.text
    payload = own_json.json()
    assert payload["metadata"]["scope"] == "SELF"
    assert payload["metadata"]["total_count"] == 1
    assert [item["email"] for item in payload["items"]] == ["employee@company.com"]
    assert all("monthly_salary" not in item for item in payload["items"])

    audit = transactional_db_session.query(AuditLog).filter(
        AuditLog.tool_name == "export_hr_directory"
    ).order_by(AuditLog.created_at.desc()).first()
    assert audit is not None
    assert audit.output_result["allowed_sections"] == ["BASIC"]


def test_hr_leave_balance_query(client, employee_token_headers):
    """Test HR Agent handling leave balance inquiry."""
    payload = {
        "agent_role": "HR",
        "message": "Tôi còn bao nhiêu ngày phép?",
    }
    response = client.post("/api/v1/agent/chat", json=payload, headers=employee_token_headers)
    assert response.status_code == 200
    data = response.json()
    assert data["agent_role"] == "HR"
    assert "ngày phép" in data["reply"].lower()
    assert len(data["tools_executed"]) > 0
    assert data["tools_executed"][0]["tool_name"] == "query_leave_balance"
    assert data["hr_card"]["type"] == "LEAVE_BALANCE"

    conversation = client.get(
        f"/api/v1/agent/conversations/{data['conversation_id']}",
        headers=employee_token_headers,
    )
    assert conversation.status_code == 200
    assistant = conversation.json()["messages"][-1]
    assert assistant["attachments"][0]["type"] == "HR_CARD"
    assert assistant["attachments"][0]["payload"]["type"] == "LEAVE_BALANCE"


def test_hr_leave_request_creates_approval_card(client, employee_token_headers):
    """Test HR Agent handling leave submission and generating an Approval Card."""
    # Reset leave balance memory for employee to ensure deterministic quota
    db = SyncSessionLocal()
    try:
        user = db.query(User).filter(User.email == "employee@company.com").first()
        if user:
            mem = db.query(UserMemory).filter(
                UserMemory.user_id == user.id,
                UserMemory.memory_key == "leave_balance",
            ).first()
            if mem:
                mem.memory_value = json.dumps({"total_days": 12, "used_days": 2, "remaining_days": 10})
                db.commit()
    finally:
        db.close()

    start = date.today() + timedelta(days=30)
    while start.weekday() >= 5:
        start += timedelta(days=1)
    end = start + timedelta(days=1)
    while end.weekday() >= 5:
        end += timedelta(days=1)
    payload = {
        "agent_role": "HR",
        "message": (
            f"Tôi muốn xin nghỉ phép từ {start.isoformat()} đến {end.isoformat()} "
            "vì lý do gia đình"
        ),
    }
    response = client.post("/api/v1/agent/chat", json=payload, headers=employee_token_headers)
    assert response.status_code == 200
    data = response.json()
    assert data["agent_role"] == "HR"
    assert data["approval_card"] is not None
    card = data["approval_card"]
    assert card["action_type"] == "XIN NGHỈ PHÉP"
    assert card["status"] == "WAITING"
    assert "id" in card
